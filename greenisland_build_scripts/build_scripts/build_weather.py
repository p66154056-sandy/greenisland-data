# -*- coding: utf-8 -*-
"""
產生 REAL_WEATHER.json（⑪氣象資料用）
來源：11_2015至2024年綠島自動氣象站氣象資料.xlsx（4 個工作表：月均溫及年均溫/最高溫/最低溫/降雨量）

特殊處理：
- CODiS 綠島自動氣象站 2023/1-2024/7 無氣溫資料，2023、2024 年皆使用 LTSER LYUDAO
  海洋研究站每日氣象資料代替（Excel 內已經是合併好的資料，此腳本僅負責解析格式）
- 最高溫/最低溫欄位混合兩種日期格式：
  "數值 / YYYY/MM/DD HH:MM:SS"（CODiS）與 "數值(YYYY-MM-DDTHH:MM:SS)"（LTSER）
- 2019年4-5月、2023年1/2/12月無資料（原始欄位為 "--"）
"""
import json
import re
import openpyxl
from common import UPLOAD_DIR, YEARS, OUTPUT_DIR


def parse_val_date(cell):
    if cell is None:
        return None, None
    s = str(cell).strip()
    if s in ("--", "-- / --", ""):
        return None, None
    m = re.match(r"^([\d.]+)\s*/\s*(.+)$", s)
    if m:
        val = float(m.group(1))
        dm = re.match(r"(\d{4})/(\d{2})/(\d{2})", m.group(2).strip())
        date_fmt = f"{dm.group(1)}-{dm.group(2)}-{dm.group(3)}" if dm else m.group(2).strip()
        return val, date_fmt
    m2 = re.match(r"^([\d.]+)\((.+)\)$", s)
    if m2:
        val = float(m2.group(1))
        dm2 = re.match(r"(\d{4}-\d{2}-\d{2})", m2.group(2).strip())
        date_fmt = dm2.group(1) if dm2 else m2.group(2).strip()
        return val, date_fmt
    try:
        return float(s), None
    except ValueError:
        return None, None


def sheet_to_rows(ws):
    rows = {}
    for r in range(2, 12):
        year = ws.cell(row=r, column=1).value
        if year is None:
            continue
        rows[int(year)] = [ws.cell(row=r, column=c).value for c in range(2, 15)]
    return rows


def build_weather():
    wb = openpyxl.load_workbook(f"{UPLOAD_DIR}/11_2015至2024年綠島自動氣象站氣象資料.xlsx", data_only=True)
    avg_rows = sheet_to_rows(wb["月均溫及年均溫"])
    max_rows = sheet_to_rows(wb["最高溫"])
    min_rows = sheet_to_rows(wb["最低溫"])
    rain_rows = sheet_to_rows(wb["降雨量"])

    result = {}
    for y in YEARS:
        avg = avg_rows[y]
        monthly_avg = [None if str(v).strip() == "--" else round(float(v), 1) for v in avg[:12]]
        year_avg = None if str(avg[12]).strip() == "--" else round(float(avg[12]), 1)

        monthly_max, monthly_min = [], []
        for v in max_rows[y][:12]:
            val, _ = parse_val_date(v)
            monthly_max.append(round(val, 1) if val is not None else None)
        for v in min_rows[y][:12]:
            val, _ = parse_val_date(v)
            monthly_min.append(round(val, 1) if val is not None else None)

        y_max_val, y_max_date = parse_val_date(max_rows[y][12])
        y_min_val, y_min_date = parse_val_date(min_rows[y][12])

        rain = rain_rows[y]
        monthly_rain = [None if v is None or str(v).strip() == "--" else round(float(v), 1) for v in rain[:12]]
        year_rain = None if rain[12] is None else round(float(rain[12]), 1)

        result[str(y)] = {
            "monthlyAvg": monthly_avg, "monthlyMax": monthly_max, "monthlyMin": monthly_min,
            "monthlyRain": monthly_rain,
            "yearAvg": year_avg,
            "yearMaxVal": round(y_max_val, 1) if y_max_val else None, "yearMaxDate": y_max_date,
            "yearMinVal": round(y_min_val, 1) if y_min_val else None, "yearMinDate": y_min_date,
            "yearRain": year_rain,
        }

    with open(f"{OUTPUT_DIR}/REAL_WEATHER.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1)
    print("REAL_WEATHER.json 完成")
    return result


if __name__ == "__main__":
    build_weather()
